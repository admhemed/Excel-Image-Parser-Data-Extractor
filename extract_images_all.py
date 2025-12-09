import os
from uuid import uuid4
from collections import Counter
from typing import List, Dict, Any, Tuple, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage

# ==========================
# إعدادات عامة
# ==========================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# الروت المستهدف الذي يحتوي ملفات الإكسل
ROOT_DIR = os.path.join(SCRIPT_DIR, "TurnKeyFiles2")

# فولدر الصور (بجانب السكربت)
IMAGES_DIR = os.path.join(SCRIPT_DIR, "images")

# اسم ملف الداتا الناتج (بجانب السكربت أيضاً)
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "packages_data.xlsx")

# كلمات الهيدر التي نبحث عنها (تُقارن بحروف صغيرة)
HEADER_KEYWORDS = ("part number", "description", "qty")

# قيم إكسل الخطأ التي لا يجب اعتبارها أسماء باكجات
EXCLUDED_PACKAGE_TOKENS = {
    "#unknown!",
    "#value!",
    "#div/0!",
    "#ref!",
    "#name?",
    "#null!",
    "#num!",
    "#n/a",
}

# أكواد ألوان ANSI للّوغ (بدون مكتبات إضافية)
RESET = "\033[0m"
CYAN = "\033[36m"
YELLOW = "\033[33m"
GREEN = "\033[32m"
RED = "\033[31m"
MAGENTA = "\033[35m"


def log_info(msg: str) -> None:
    print(f"{CYAN}[INFO]{RESET} {msg}")


def log_warn(msg: str) -> None:
    print(f"{YELLOW}[WARN]{RESET} {msg}")


def log_success(msg: str) -> None:
    print(f"{GREEN}[OK]{RESET}   {msg}")


def log_error(msg: str) -> None:
    print(f"{RED}[ERROR]{RESET} {msg}")


def log_debug(msg: str) -> None:
    print(f"{MAGENTA}[DEBUG]{RESET} {msg}")


# ==========================
# توابع مساعدة عامة
# ==========================

def to_int_or_none(value) -> Optional[int]:
    """
    محاولة تحويل القيمة إلى int وإرجاع None لو فشل التحويل أو كانت القيمة فارغة.
    مشابهة تماماً لما كان يحدث في نسخة pandas.
    """
    if value is None:
        return None
    try:
        text = str(value).strip()
        if not text:
            return None
        # أحياناً تكون القيمة float مثل 3.0
        f = float(text)
        return int(f)
    except (ValueError, TypeError):
        return None


# ==========================
# إدارة فولدر الصور
# ==========================

def ensure_clean_images_dir() -> None:
    """
    يتأكد من وجود فولدر الصور ويمسح أي ملفات موجودة بداخله قبل البدء.
    يعمل داخل IMAGES_DIR.
    """
    if os.path.isdir(IMAGES_DIR):
        for name in os.listdir(IMAGES_DIR):
            path = os.path.join(IMAGES_DIR, name)
            if os.path.isfile(path):
                os.remove(path)
    else:
        os.makedirs(IMAGES_DIR, exist_ok=True)

    log_info(f"Images directory ready and cleaned: '{IMAGES_DIR}'")


def get_image_bytes(img) -> Optional[bytes]:
    """
    تحاول استخراج بايتات الصورة من كائن openpyxl Image.
    نعتمد على الخاصية _data (قد تكون دالة أو بايتات جاهزة).
    """
    data_attr = getattr(img, "_data", None)

    if callable(data_attr):
        try:
            return data_attr()
        except Exception as e:
            log_warn(f"Failed to call img._data(): {e}")

    if isinstance(data_attr, (bytes, bytearray)):
        return bytes(data_attr)

    log_warn("Could not extract image bytes from Image object.")
    return None


def guess_image_ext(data: bytes) -> str:
    """
    تخمين بسيط لامتداد الصورة من أول بايتات بدون استخدام مكتبات إضافية.
    لو ما عرفنا النوع نرجّع 'jpg' افتراضياً.
    """
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if data.startswith(b"\xff\xd8"):
        return "jpg"
    if data.startswith(b"GIF87a") or data.startswith(b"GIF89a"):
        return "gif"
    if data.startswith(b"BM"):
        return "bmp"
    return "jpg"


# ==========================
# حساب إحداثيات Y للصفوف
# ==========================

def compute_row_y_map(ws):
    """
    تحسب إحداثيات Y (بالـ EMU) لكل صف:
    top_y[row] = موضع بداية الصف من الأعلى
    bottom_y[row] = موضع نهاية الصف
    نعتمد على ارتفاع الصفوف (إن كان مخصصاً) أو ارتفاع الديفولت من sheet_format.
    """
    sheet_format = ws.sheet_format
    default_height_points = sheet_format.defaultRowHeight or 15  # نقاط
    EMU_PER_POINT = 12700  # ثابت تحويل من نقاط إلى EMU

    top_y = {}
    bottom_y = {}
    acc = 0

    for r in range(1, ws.max_row + 1):
        h = ws.row_dimensions[r].height
        if h is None:
            h = default_height_points
        top_y[r] = acc
        acc += h * EMU_PER_POINT
        bottom_y[r] = acc

    log_info(
        f"Row Y mapping computed using default height {default_height_points} pt "
        f"(~{default_height_points * EMU_PER_POINT:.0f} EMU per row when not custom)."
    )
    return top_y, bottom_y


# ==========================
# منطق استخراج الباكجات
# ==========================

def find_first_header_row(ws) -> int:
    """
    تبحث عن أول سطر يحتوي على أي من الكلمات:
    Part Number / Description / Qty
    في أي عمود من الأعمدة، وترجع رقم السطر (1-based).
    لو لم تجده ترجع 0.
    """
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value:
                continue

            text = str(cell_value).strip().lower()
            if any(keyword in text for keyword in HEADER_KEYWORDS):
                log_info(
                    f"First header row detected at row {row} "
                    f"(col {col}) with value '{cell_value}'"
                )
                return row

    log_warn("No header row found with Part Number / Description / Qty in any column.")
    return 0


def build_packages(ws, top_y, bottom_y) -> List[Dict[str, Any]]:
    """
    تبني قائمة الباكجات اعتماداً على العمود الأول.
    نفس المنطق السابق مع y_start و y_end.
    """
    header_row = find_first_header_row(ws)
    if header_row <= 1:
        log_error("Cannot determine package start row (header row not found or at first row).")
        return []

    start_row = header_row - 1
    log_info(f"Package detection will start from row {start_row} (row above first header).")

    packages: List[Dict[str, Any]] = []
    current_package: Optional[Dict[str, Any]] = None

    for row in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None:
            continue

        text = str(cell_value).strip()
        lower_text = text.lower()

        if any(keyword in lower_text for keyword in HEADER_KEYWORDS):
            continue

        if lower_text in EXCLUDED_PACKAGE_TOKENS:
            log_debug(f"Ignoring error-like value at row {row}: '{text}'")
            continue

        if current_package is not None:
            current_package["end_row"] = row - 1
            current_package["y_start"] = top_y[current_package["start_row"]]
            current_package["y_end"] = bottom_y[current_package["end_row"]]
            packages.append(current_package)

        current_package = {
            "name": text,
            "start_row": row,
            "end_row": None,
            "y_start": None,
            "y_end": None,
            "images": [],
            "abs_images": [],
            "uid": None,
            "image_filename": None,
            "category": None,
        }
        # log_debug(f"New package detected at row {row}: '{text}'")

    if current_package is not None:
        current_package["end_row"] = ws.max_row
        current_package["y_start"] = top_y[current_package["start_row"]]
        current_package["y_end"] = bottom_y[current_package["end_row"]]
        packages.append(current_package)

    if not packages:
        log_warn("No packages were detected.")
        return []

    log_success(f"{len(packages)} packages detected.")
    return packages


def fill_packages_categories(ws, packages: List[Dict[str, Any]]) -> None:
    """
    ملء Category من العمود F ضمن مدى أسطر كل باكج.
    """
    CATEGORY_COL = 6  # العمود F

    for pkg in packages:
        category = None

        for row in range(pkg["start_row"], pkg["end_row"] + 1):
            cell_value = ws.cell(row=row, column=CATEGORY_COL).value
            if isinstance(cell_value, str):
                text = cell_value.strip()
                if text:
                    category = text
                    break

        pkg["category"] = category
        log_debug(
            f"Package '{pkg['name']}' rows [{pkg['start_row']}-{pkg['end_row']}]: "
            f"Category = '{category}'"
        )


# ==========================
# اكتشاف أعمدة No / PartNo / Desc / QTY
# ==========================

def detect_detail_columns(ws, first_package: Dict[str, Any]) -> Optional[Tuple[int, int, int, int]]:
    """
    نبحث عن عمود رقم السطر (No) بهذه الأولوية:
    1) عمود عنوانه بالضبط "#"
    2) أو عمود عنوانه "no" / "No" / "No." الخ
    ثم نعتبر ما بعده: Part, Desc, QTY.
    """

    start_row = first_package["start_row"]
    end_row = first_package["end_row"]

    FIRST_DATA_COL = 2
    MAX_OFFSET_COLS = 5
    last_col_to_check = FIRST_DATA_COL + MAX_OFFSET_COLS - 1

    candidate_no_cols = []

    for row in range(start_row, end_row + 1):
        for col in range(FIRST_DATA_COL, last_col_to_check + 1):
            cell_value = ws.cell(row=row, column=col).value
            if not isinstance(cell_value, str):
                continue

            raw = cell_value.strip()
            lower = raw.lower()

            # الشرط الجديد: إما بالضبط "#" أو "no" (مع بعض الأشكال البسيطة)
            is_hash = (raw == "#")
            is_no   = (lower in {"no", "no.", "no#", "no:"})

            if not (is_hash or is_no):
                continue

            candidate_no_cols.append((row, col, raw))

    if not candidate_no_cols:
        log_warn(
            "Could not detect detail columns (No / Part Number / Description / QTY) "
            "inside the first package range."
        )
        return None

    # لو وجدنا أكثر من واحد، نختار أول واحد (أقرب شيء للأعلى)
    row, col_no, header_text = candidate_no_cols[0]

    col_part = col_no + 1
    col_desc = col_no + 2
    col_qty  = col_no + 3

    if col_qty > ws.max_column:
        log_warn(
            f"Detected No-like header '{header_text}' at col {col_no} row {row} "
            f"but following columns exceed max_column={ws.max_column}."
        )
        return None

    log_info(
        f"Detail columns detected (row {row}, header '{header_text}'): "
        f"No={col_no}, PartNo={col_part}, Desc={col_desc}, QTY={col_qty}"
    )
    return col_no, col_part, col_desc, col_qty


# ==========================
# ربط الصور بالباكجات
# ==========================

def find_package_for_row(packages: List[Dict[str, Any]], row: int) -> Optional[Dict[str, Any]]:
    for pkg in packages:
        if pkg["start_row"] <= row <= pkg["end_row"]:
            return pkg
    return None


def find_package_for_y_center(packages: List[Dict[str, Any]], center_y: int) -> Optional[Dict[str, Any]]:
    for pkg in packages:
        if pkg["y_start"] is None or pkg["y_end"] is None:
            continue
        if pkg["y_start"] <= center_y < pkg["y_end"]:
            return pkg
    return None


def collect_worksheet_images(ws):
    images = getattr(ws, "_images", [])
    log_info(f"Total images found: {len(images)}")
    log_info(f"Anchor types count: {Counter(type(img.anchor).__name__ for img in images)}")
    return images


def map_images_to_packages(images, packages: List[Dict[str, Any]]) -> List[int]:
    unmatched_images: List[int] = []

    for idx, img in enumerate(images):
        anchor = img.anchor
        tname = type(anchor).__name__

        if tname in ["OneCellAnchor", "TwoCellAnchor"] and hasattr(anchor, "_from") and anchor._from is not None:
            fm = anchor._from
            row_zero_based = getattr(fm, "row", 0)
            row_excel = row_zero_based + 1

            pkg = find_package_for_row(packages, row_excel)
            if pkg:
                if pkg["images"] or pkg["abs_images"]:
                    log_warn(
                        f"Ignoring extra image #{idx} for package '{pkg['name']}' "
                        f"(already has an image)."
                    )
                else:
                    pkg["images"].append(idx)
                    log_success(
                        f"Image #{idx} (OneCellAnchor at row {row_excel}) linked to package "
                        f"'{pkg['name']}' [rows {pkg['start_row']} - {pkg['end_row']}]"
                    )
            else:
                unmatched_images.append(idx)
                log_warn(
                    f"Image #{idx} (OneCellAnchor at row {row_excel}) "
                    f"could not be matched to any package."
                )
            continue

        if tname == "AbsoluteAnchor" and hasattr(anchor, "pos") and anchor.pos is not None:
            pos = anchor.pos
            ext = getattr(anchor, "ext", None)

            if ext is None:
                unmatched_images.append(idx)
                log_warn(
                    f"Image #{idx} (AbsoluteAnchor) has no ext; skipped for package mapping."
                )
                continue

            y_top = getattr(pos, "y", None)
            cy = getattr(ext, "cy", None)

            if y_top is None or cy is None:
                unmatched_images.append(idx)
                log_warn(
                    f"Image #{idx} (AbsoluteAnchor) missing pos.y or ext.cy; skipped for mapping."
                )
                continue

            center_y = y_top + cy / 2
            pkg = find_package_for_y_center(packages, center_y)

            if pkg:
                if pkg["images"] or pkg["abs_images"]:
                    log_warn(
                        f"Ignoring extra image #{idx} for package '{pkg['name']}' "
                        f"(already has an image)."
                    )
                else:
                    pkg["abs_images"].append(idx)
                    log_success(
                        f"Image #{idx} (AbsoluteAnchor center_y={center_y:.0f}) linked to package "
                        f"'{pkg['name']}' [Y {pkg['y_start']:.0f} - {pkg['y_end']:.0f}]"
                    )
            else:
                unmatched_images.append(idx)
                log_warn(
                    f"Image #{idx} (AbsoluteAnchor center_y={center_y:.0f}) "
                    f"could not be matched to any package."
                )
            continue

        unmatched_images.append(idx)
        log_warn(f"Image #{idx} with anchor type '{tname}' could not be processed for mapping.")

    return unmatched_images


def assign_uids_and_save_images(images, packages: List[Dict[str, Any]]) -> None:
    # log_info("=== Package list (id + optional image) ===")
    # print("package_name\tstart_row\tid\timage")

    for pkg in packages:
        img_idx = None
        if pkg["images"]:
            img_idx = pkg["images"][0]
        elif pkg["abs_images"]:
            img_idx = pkg["abs_images"][0]

        uid = str(uuid4())
        filename = None

        if img_idx is not None:
            img_obj = images[img_idx]
            img_bytes = get_image_bytes(img_obj)

            if not img_bytes:
                log_warn(
                    f"Image bytes for image #{img_idx} (package '{pkg['name']}') "
                    f"could not be extracted."
                )
            else:
                ext = guess_image_ext(img_bytes)
                filename = f"{uid}.{ext}"
                filepath = os.path.join(IMAGES_DIR, filename)

                try:
                    with open(filepath, "wb") as f:
                        f.write(img_bytes)

                    # log_success(
                    #     f"Saved image #{img_idx} for package '{pkg['name']}' "
                    #     f"as '{filename}'"
                    # )
                except Exception as e:
                    log_error(f"Failed to save image for package '{pkg['name']}': {e}")
                    filename = None

        pkg["uid"] = uid
        pkg["image_filename"] = filename

        # print(f"{pkg['name']}\t{pkg['start_row']}\t{uid}\t{filename or ''}")


def link_images_to_packages(ws, packages: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    images = collect_worksheet_images(ws)
    unmatched_images = map_images_to_packages(images, packages)
    assign_uids_and_save_images(images, packages)

    if unmatched_images:
        log_warn(f"Unmatched images: {unmatched_images}")

    return packages


# ==========================
# فك الدمج العمودي (forward-fill) في أعمدة التفاصيل
# ==========================
def flatten_vertical_merges_in_column(ws, col: int) -> None:
    """
    يفك الدمج العمودي في عمود واحد (مثل عمود No أو QTY):
    - يبحث في ws.merged_cells عن أي range من نوع عمودي في هذا العمود (min_col == max_col == col)
    - يأخذ قيمة الخلية الأولى (أعلى سطر في الدمج)
    - يعمل unmerge للـ range
    - يكتب نفس القيمة في كل الأسطر ضمن هذا الدمج لهذا العمود
    """
    # نأخذ نسخة من القائمة لأننا سنعدل الـ merged_cells أثناء الدوران
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        # نهتم فقط بحالات الدمج العمودي في هذا العمود بالذات
        if min_col == max_col == col and max_row > min_row:
            # قيمة الخلية الأصلية (أعلى الخلية في الدمج)
            value = ws.cell(row=min_row, column=col).value

            # نفك الدمج
            ws.unmerge_cells(str(merged_range))

            # ننسخ القيمة على كل الأسطر في هذا العمود
            if value is not None and str(value).strip() != "":
                for row in range(min_row, max_row + 1):
                    ws.cell(row=row, column=col).value = value

            log_debug(
                f"Flattened vertical merge in col {col} "
                f"rows [{min_row}-{max_row}] with value '{value}'"
            )


def forward_fill_column_in_range(ws, col: int, start_row: int, end_row: int) -> None:
    last_value = None

    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        value = cell.value

        if value is not None and str(value).strip() != "":
            last_value = value
        else:
            if last_value is not None:
                cell.value = last_value


def find_data_rows_range_for_package(
    ws,
    pkg: Dict[str, Any],
    col_part: int,
    col_desc: int,
) -> Optional[Tuple[int, int]]:
    """
    يحدد نطاق أسطر البيانات (parts) لكل باكج:
    - أي سطر فيه Part Number أو Description نعتبره داتا.
    - يرجع (data_start, data_end) أو None لو لم يوجد داتا.
    هذا يحاكي منطق data_rows في كود pandas القديم.
    """
    data_start = None
    data_end = None

    for row in range(pkg["start_row"], pkg["end_row"] + 1):
        part_val = ws.cell(row=row, column=col_part).value
        desc_val = ws.cell(row=row, column=col_desc).value

        has_part = (
            part_val is not None and str(part_val).strip() != ""
        )
        has_desc = (
            desc_val is not None and str(desc_val).strip() != ""
        )

        if has_part or has_desc:
            if data_start is None:
                data_start = row
            data_end = row

    if data_start is None or data_end is None:
        return None

    return data_start, data_end


def normalize_merged_detail_cells_for_all_packages(
    ws,
    packages: List[Dict[str, Any]],
    col_no: int,
    col_part: int,
    col_desc: int,
    col_qty: int,
) -> Dict[int, Tuple[int, int]]:
    """
    تعوّض الدمج العمودي داخل أعمدة تفاصيل القطع ضمن مجال أسطر الداتا لكل باكج.
    ترجع dict يربط package_index → (data_start, data_end) لاستخدامه لاحقاً.
    """

    # أولاً: نفك الدمج العمودي في عمودي No و QTY على مستوى الشيت كله
    # (لأن نفس الدمج قد يمر بعدة باكجات، وأسهل نفكه مرة واحدة)
    flatten_vertical_merges_in_column(ws, col_no)
    flatten_vertical_merges_in_column(ws, col_qty)

    data_ranges: Dict[int, Tuple[int, int]] = {}

    for idx, pkg in enumerate(packages):
        res = find_data_rows_range_for_package(ws, pkg, col_part, col_desc)
        if res is None:
            log_warn(f"No data rows found for package '{pkg['name']}'.")
            continue

        data_start, data_end = res
        data_ranges[idx] = (data_start, data_end)

        # بعد فك الدمج، نعمل forward-fill ضمن نطاق بيانات القطع فقط
        forward_fill_column_in_range(ws, col_no, data_start, data_end)
        forward_fill_column_in_range(ws, col_qty, data_start, data_end)

        log_debug(
            f"Forward-filled No/QTY for package '{pkg['name']}' "
            f"rows [{data_start}-{data_end}]."
        )

    return data_ranges


# ==========================
# معالجة ملف واحد واستخراج أسطر القطع
# ==========================

def process_workbook(path: str) -> List[tuple]:
    """
    تعالج ملف إكسل واحد:
    - تبني الباكجات + الفئات + الصور + الأعمدة التفصيلية + فك الدمج
    - تبني أسطر القطع لكل باكج:
      (PackageId, ImagePath, TitleTrim, PackageName,
       No, PartNo, PartNameAndStandard, QTY, Category)
    """
    basename = os.path.basename(path)
    title_trim = os.path.splitext(basename)[0].strip()

    log_info(f"Opening workbook: {basename}")

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        log_error(f"Failed to open '{basename}': {e}")
        return []

    ws = wb.active

    top_y, bottom_y = compute_row_y_map(ws)

    packages = build_packages(ws, top_y, bottom_y)
    if not packages:
        log_warn(f"No packages found in '{basename}'.")
        return []

    fill_packages_categories(ws, packages)

    link_images_to_packages(ws, packages)

    detail_cols = detect_detail_columns(ws, packages[0])
    if detail_cols is None:
        log_warn(
            f"Detail columns could not be detected in '{basename}'. "
            f"Only package-level rows will be emitted without part details."
        )
        # fallback: نرجّع سطر واحد لكل باكج بدون تفاصيل
        rows_for_excel: List[tuple] = []
        for pkg in packages:
            uid = pkg.get("uid")
            image_filename = pkg.get("image_filename") or ""
            pkg_name = pkg["name"]
            category = pkg.get("category") or ""
            rows_for_excel.append(
                (uid, image_filename, title_trim, pkg_name, "", "", "", "", category)
            )
        return rows_for_excel

    col_no, col_part, col_desc, col_qty = detail_cols

    # نطبّق forward-fill على No و QTY ضمن مدى أسطر الداتا لكل باكج
    data_ranges_by_pkg_index = normalize_merged_detail_cells_for_all_packages(
        ws,
        packages,
        col_no,
        col_part,
        col_desc,
        col_qty,
    )

    rows_for_excel: List[tuple] = []

    # نمر على كل باكج ونبني أسطر القطع
    for idx, pkg in enumerate(packages):
        if idx not in data_ranges_by_pkg_index:
            # باكج بدون داتا حقيقية
            continue

        data_start, data_end = data_ranges_by_pkg_index[idx]

        uid = pkg.get("uid")
        image_filename = pkg.get("image_filename") or ""
        pkg_name = pkg["name"]
        category = pkg.get("category") or ""

        for row in range(data_start, data_end + 1):
            no_val = ws.cell(row=row, column=col_no).value
            part_val = ws.cell(row=row, column=col_part).value
            desc_val = ws.cell(row=row, column=col_desc).value
            qty_val = ws.cell(row=row, column=col_qty).value

            # نتأكد أن السطر فيه Part أو Description
            has_part = part_val is not None and str(part_val).strip() != ""
            has_desc = desc_val is not None and str(desc_val).strip() != ""
            if not (has_part or has_desc):
                continue

            # رقم الـ No يجب أن يكون عدداً صحيحاً (مثل النسخة القديمة)
            no_int = to_int_or_none(no_val)
            if no_int is None:
                continue

            qty_int = to_int_or_none(qty_val)

            part_str = str(part_val).strip() if part_val is not None else ""
            desc_str = str(desc_val).strip() if desc_val is not None else ""

            rows_for_excel.append(
                (
                    uid,            # PackageId
                    image_filename, # ImagePath
                    title_trim,     # Title - TRIM
                    pkg_name,       # PackageName
                    no_int,         # No
                    part_str,       # PartNo
                    desc_str,       # Part Name And Standard
                    qty_int,        # QTY
                    category,       # Category
                )
            )

    return rows_for_excel


# ==========================
# الدالة الرئيسية
# ==========================
def main():
    """
    - يحضّر فولدر الصور.
    - يعالج كل ملف .xlsx في ROOT_DIR وفي المجلدات الفرعية داخله.
    - يبني ملف إكسل جديد packages_data.xlsx
      يحتوي أسطر القطع مع تكرار بيانات الباكج لكل سطر.
    - يضع الصورة في عمود مستقل (B) واسم الملف في عمود مستقل (C).
    - يترك سطرًا فارغًا بعد أسطر كل باكج قبل بدء الباكج التالية.
    """
    ensure_clean_images_dir()

    # نجمع كل ملفات .xlsx في ROOT_DIR وفي كل المجلدات الفرعية
    xlsx_files: List[str] = []
    for dirpath, _, filenames in os.walk(ROOT_DIR):
        for name in filenames:
            lower = name.lower()
            if not lower.endswith(".xlsx"):
                continue
            if name.startswith("~$"):
                # ملفات مؤقتة لإكسل
                continue
            xlsx_files.append(os.path.join(dirpath, name))

    if not xlsx_files:
        log_error(f"No .xlsx files found under ROOT_DIR: {ROOT_DIR}")
        return

    log_info(f"Found {len(xlsx_files)} .xlsx file(s) under ROOT_DIR: {ROOT_DIR}")
    for path in xlsx_files:
        rel = os.path.relpath(path, ROOT_DIR)
        log_debug(f"- {rel}")

    all_rows: List[tuple] = []

    # معالجة كل ملف إكسل وجمع كل أسطر الداتا
    for full_path in xlsx_files:
        print()
        rel = os.path.relpath(full_path, ROOT_DIR)
        print(f"{MAGENTA}========== Processing file: {rel} =========={RESET}")
        rows = process_workbook(full_path)
        all_rows.extend(rows)

    if not all_rows:
        log_warn("No rows were collected. Excel data file will not be created.")
        return

    # إنشاء ملف الإكسل الناتج
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "packages"

    # عرض الأعمدة
    ws_out.column_dimensions["A"].width = 40
    ws_out.column_dimensions["B"].width = 9
    ws_out.column_dimensions["C"].width = 10
    ws_out.column_dimensions["D"].width = 30
    ws_out.column_dimensions["E"].width = 30
    ws_out.column_dimensions["F"].width = 5
    ws_out.column_dimensions["G"].width = 18
    ws_out.column_dimensions["H"].width = 20
    ws_out.column_dimensions["I"].width = 4


    # الهيدر
    header = [
        "PackageId",
        "Image",       # عمود معاينة الصورة
        "ImagePath",   # اسم ملف الصورة (نصي فقط)
        "Title - TRIM",
        "PackageName",
        "No",
        "PartNo",
        "Part Name And Standard",
        "QTY",
        "Category",
        "delete",
        "price",
        "Description",
        "Old Part No.",
        "Names and specifications of old parts",
        "note",
        "is_red",
        "is_line",
        "is_deleted",
        "is_orange",
        "is_pink",
        "is_yellow",
        "internal_notes",
    ]
    ws_out.append(header)

    # عدد الأعمدة (نستخدمه عندما نضيف سطر فارغ)
    num_cols = len(header)

    # الهيدر في الصف 1، نبدأ العد من هناك
    row_idx = 1
    last_pkg_key = None  # لتتبع تغيّر الباكج

    # البيانات + إدراج الصور مع سطر فارغ بين كل باكج والتي تليها
    for row_data in all_rows:
        (
            uid,
            filename,
            title_trim,
            pkg_name,
            no_val,
            part_no,
            part_name_std,
            qty_val,
            category,
        ) = row_data

        # تعريف الباكج: uid + عنوان الملف + اسم الباكج
        pkg_key = (uid, title_trim, pkg_name)

        # لو تغيّرت الباكج عن السابقة → نضيف سطر فارغ
        if last_pkg_key is not None and pkg_key != last_pkg_key:
            row_idx += 1
            ws_out.append([""] * num_cols)  # سطر فارغ تماماً

        # نضيف سطر الداتا لهذه الباكج
        row_idx += 1
        ws_out.append([
            uid,           # PackageId
            "",            # Image (الصورة فقط، لا نص)
            filename,      # ImagePath
            title_trim,    # Title - TRIM
            pkg_name,      # PackageName
            no_val,        # No
            part_no,       # PartNo
            part_name_std, # Part Name And Standard
            qty_val,       # QTY
            category,      # Category
            "",            # delete
            "",            # price
            "",            # Description
            "",            # Old Part No.
            "",            # Names and specifications of old parts
            "",            # note
            "",            # is_red
            "",            # is_line
            "",            # is_deleted
            "",            # is_orange
            "",            # is_pink
            "",            # is_yellow
            "",            # internal_notes
        ])

        # إدراج الصورة في العمود B لنفس الصف
        if filename:
            img_path = os.path.join(IMAGES_DIR, filename)
            if os.path.exists(img_path):
                try:
                    xl_img = XLImage(img_path)
                    xl_img.width = 50
                    xl_img.height = 50
                    ws_out.add_image(xl_img, f"B{row_idx}")
                    ws_out.row_dimensions[row_idx].height = 35
                except Exception as e:
                    log_warn(f"Failed to embed image '{img_path}' into Excel: {e}")

        last_pkg_key = pkg_key

    wb_out.save(OUTPUT_EXCEL)
    log_success(f"Data Excel file created: '{OUTPUT_EXCEL}'")

if __name__ == "__main__":
    main()
