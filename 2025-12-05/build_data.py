import os
import uuid
import shutil
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from PIL import Image as PILImage

# اسم ملف الخرج
OUTPUT_FILENAME = "generated_data.xlsx"
# اسم فولدر الصور تحت الروت
IMAGES_FOLDER_NAME = "images"

# ثابت للتحويل من point إلى EMU (وحدة القياس الداخلية في إكسل)
EMU_PER_POINT = 12700


def find_manual_files(root_dir: str):
    """ابحث عن كل ملفات الـ xlsx (ما عدا ملف الداتا الناتج) داخل الروت."""
    manual_files = []
    for dirpath, _, filenames in os.walk(root_dir):
        for name in filenames:
            lower = name.lower()
            if not lower.endswith(".xlsx"):
                continue
            if lower.startswith("~$"):
                # ملفات مؤقتة يفتحها إكسل
                continue
            if lower == OUTPUT_FILENAME.lower():
                # لا نعيد معالجة ملف الداتا الناتج
                continue
            manual_files.append(os.path.join(dirpath, name))
    return manual_files


def to_int_or_none(value):
    """محاولة تحويل القيمة إلى int وإرجاع None لو فشل."""
    try:
        if pd.isna(value):
            return None
        return int(float(value))
    except (ValueError, TypeError):
        return None


def build_row_boundaries(ws):
    """
    نبني حدود تقريبية لكل صف في الشيت:
    لكل صف نعطي (row_index_0_based, start_y, end_y) بوحدة EMU.
    هذا مفيد عندما تكون الصورة من نوع AbsoluteAnchor فيها pos.y بدون row.
    """
    boundaries = []
    y = 0

    default_height_pts = getattr(ws.sheet_format, "defaultRowHeight", None)
    if default_height_pts is None:
        default_height_pts = 15  # ارتفاع افتراضي معقول

    for r in range(1, ws.max_row + 1):
        dim = ws.row_dimensions.get(r)
        if dim is not None and dim.height is not None:
            h_pts = dim.height
        else:
            h_pts = default_height_pts

        h_emu = h_pts * EMU_PER_POINT
        start_y = y
        end_y = y + h_emu

        # r-1 لأننا نستخدم 0-based مثل pandas
        boundaries.append((r - 1, start_y, end_y))
        y = end_y

    return boundaries


def approx_row_from_y(boundaries, y_pos):
    """
    لو عندنا y_pos (بوحدة EMU) بدون row،
    نحاول إيجاد أقرب صف له اعتماداً على centers للـ boundaries.
    """
    if y_pos is None or not boundaries:
        return None

    closest_row = boundaries[-1][0]
    min_dist = float("inf")

    for row_idx, start_y, end_y in boundaries:
        center = (start_y + end_y) / 2
        dist = abs(center - y_pos)
        if dist < min_dist:
            min_dist = dist
            closest_row = row_idx

    return closest_row


def extract_images(ws, images_dir: str):
    """
    تحفظ الصور من الشيت في فولدر images_dir
    وتعيد list من الدكتات:
    {
        "guid":  str,
        "filename": str,
        "row_idx": int (0-based) أو None لو فشل تحديد الصف
    }
    """
    os.makedirs(images_dir, exist_ok=True)

    boundaries = build_row_boundaries(ws)
    images = []

    for img in ws._images:
        # ---------------- حفظ الصورة على الديسك ----------------
        try:
            raw = img._data()
            pil_img = PILImage.open(BytesIO(raw))

            fmt = (pil_img.format or "JPEG").lower()
            ext = "jpg" if fmt in ("jpeg", "jpg") else fmt
        except Exception:
            # لو صار أي خطأ في قراءة الفورمات نستعمل jpg افتراضياً
            pil_img = None
            ext = "jpg"

        guid = str(uuid.uuid4())
        filename = f"{guid}.{ext}"
        full_path = os.path.join(images_dir, filename)

        try:
            if pil_img is not None:
                pil_img.save(full_path)
        except Exception:
            # لو فشل الحفظ لأي سبب نتجاهل الصورة ولا نكسر السكريبت
            pass

        # ---------------- تحديد صف الصورة ----------------
        anchor = img.anchor
        row_idx = None
        y_pos = None

        try:
            if isinstance(anchor, str):
                # مثل "A10"
                r, _ = coordinate_to_tuple(anchor)
                row_idx = r - 1
            else:
                # OneCellAnchor أو TwoCellAnchor أو AbsoluteAnchor
                if hasattr(anchor, "_from") and anchor._from is not None:
                    row_idx = anchor._from.row - 1
                elif hasattr(anchor, "from_") and anchor.from_ is not None:
                    row_idx = anchor.from_.row - 1

                # بعض الأنواع يكون فيها pos.y
                if hasattr(anchor, "pos") and anchor.pos is not None:
                    y_pos = anchor.pos.y
        except Exception:
            # نتجاهل أي خطأ ونكمل
            pass

        if row_idx is None and y_pos is not None:
            row_idx = approx_row_from_y(boundaries, y_pos)

        images.append(
            {
                "guid": guid,
                "filename": filename,
                "row_idx": row_idx,
                "used": False,  # فلاج لكي لا تستعمل الصورة لأكثر من باكيج
            }
        )

    # نرتب الصور حسب الصف (من الأعلى للأسفل)
    images.sort(key=lambda im: (im["row_idx"] is None, im["row_idx"]))
    return images


def detect_columns_and_headers(df: pd.DataFrame):
    """
    نكتشف:
    - عمود Part Number
    - عمود No
    - عمود Description
    - عمود QTY
    - أسطر الهيدر (التي فيها كلمة Part Number)

    نعتمد على كلمة "Part Number" كما طلبت.
    """
    pn_col = None
    no_col = None
    desc_col = None
    qty_col = None
    header_rows = []

    for r in df.index:
        for c in df.columns:
            v = df.iloc[r, c]
            if not isinstance(v, str):
                continue
            txt = v.strip().lower()
            if txt == "part number":
                header_rows.append(r)
                if pn_col is None:
                    pn_col = c
                    # نكتشف الأعمدة الأخرى من نفس سطر الهيدر
                    for cc in df.columns:
                        vv = df.iloc[r, cc]
                        if not isinstance(vv, str):
                            continue
                        t2 = vv.strip().lower()
                        if t2.startswith("no"):
                            no_col = cc
                        elif "description" in t2:
                            desc_col = cc
                        elif t2.startswith("qty"):
                            qty_col = cc
                break

    # لو وجدنا عمود Part Number ولم نجد الأعمدة الأخرى، نستخدم مواقع افتراضية نسبية
    if pn_col is not None:
        if no_col is None:
            no_col = pn_col - 1
        if desc_col is None:
            desc_col = pn_col + 1
        if qty_col is None:
            qty_col = pn_col + 2

    return pn_col, no_col, desc_col, qty_col, header_rows


def extract_manual_to_flat(manual_path: str, root_dir: str) -> pd.DataFrame:
    print(f"\nProcessing file: {manual_path}")

    title = os.path.splitext(os.path.basename(manual_path))[0].strip()

    # نقرأ الشيت الأولى كـ DataFrame
    xls = pd.ExcelFile(manual_path)
    sheet_name = xls.sheet_names[0]
    df = xls.parse(sheet_name, header=None)

    # نفتح نفس الشيت كـ openpyxl للحصول على الصور
    wb = load_workbook(manual_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    images_dir = os.path.join(root_dir, IMAGES_FOLDER_NAME)
    images_info = extract_images(ws, images_dir)

    # نكتشف الأعمدة والهيدر
    pn_col, no_col, desc_col, qty_col, header_rows = detect_columns_and_headers(df)
    if pn_col is None or not header_rows:
        print("  ⚠ No 'Part Number' header found. Skipping file.")
        return pd.DataFrame([])

    # ---------------- تعريف الباكيجات (sections) ----------------
    sections = []

    for i, header_row in enumerate(header_rows):
        title_row = header_row - 1  # السطر الذي يحتوي اسم الباكج مثل REAR WHEELS

        # نهاية نطاق البحث عن بيانات هذه الباكج: قبل هيدر الباكج التالية
        next_header = header_rows[i + 1] if i + 1 < len(header_rows) else len(df)

        data_rows = []
        for r in range(header_row + 1, next_header):
            if r not in df.index:
                continue
            partno = df.iloc[r, pn_col] if pn_col in df.columns else None
            desc = df.iloc[r, desc_col] if desc_col in df.columns else None

            has_part = (
                (isinstance(partno, str) and partno.strip() != "")
                or (not isinstance(partno, str) and not pd.isna(partno))
            )
            has_desc = (
                (isinstance(desc, str) and desc.strip() != "")
                or (not isinstance(desc, str) and not pd.isna(desc))
            )

            # أي سطر فيه Part Number أو Description نعتبره سطر داتا
            if has_part or has_desc:
                data_rows.append(r)

        if not data_rows:
            # باكيج بدون عناصر حقيقية، نتجاهلها
            continue

        data_start = data_rows[0]
        data_end = data_rows[-1]

        # span_start: من سطر العنوان (في العمود A) إلى آخر سطر داتا
        span_start = title_row if title_row >= 0 else data_start
        span_end = data_end

        sections.append(
            {
                "idx": i,
                "title_row": title_row,
                "header_row": header_row,
                "data_start": data_start,
                "data_end": data_end,
                "span_start": span_start,
                "span_end": span_end,
                "package_guid": None,
                "image_filename": "",
            }
        )

    # ---------------- ربط الصور بالباكيجات ----------------
    for sec in sections:
        span_start = sec["span_start"]
        span_end = sec["span_end"]

        chosen_image = None
        for img in images_info:
            if img["used"]:
                continue
            row_idx = img["row_idx"]
            if row_idx is None:
                continue
            if span_start <= row_idx <= span_end:
                chosen_image = img
                img["used"] = True
                break

        if chosen_image is not None:
            sec["package_guid"] = chosen_image["guid"]
            sec["image_filename"] = chosen_image["filename"]
        else:
            # باكيج بدون صورة، نعطيها GUID جديد لكن بدون ملف صورة
            sec["package_guid"] = str(uuid.uuid4())
            sec["image_filename"] = ""

    # ---------------- تحويل الباكيجات إلى داتا فلات ----------------
    all_rows = []

    for sec in sections:
        title_row = sec["title_row"]
        data_start = sec["data_start"]
        data_end = sec["data_end"]

        package_guid = sec["package_guid"]
        image_filename = sec["image_filename"]

        # اسم الباكج من العمود A في سطر العنوان
        package_name = ""
        if 0 in df.columns and 0 <= title_row < len(df):
            cell = df.iloc[title_row, 0]
            if isinstance(cell, str):
                package_name = cell.strip()

        # نعمل forward-fill لقيم No و QTY داخل الباكيج
        if data_start <= data_end:
            if no_col in df.columns:
                df.loc[data_start:data_end, no_col] = df.loc[data_start:data_end, no_col].ffill()
            if qty_col in df.columns:
                df.loc[data_start:data_end, qty_col] = df.loc[data_start:data_end, qty_col].ffill()

        row_idx = data_start
        while row_idx <= data_end:
            part_no = df.iloc[row_idx, pn_col] if pn_col in df.columns else None
            part_name = df.iloc[row_idx, desc_col] if desc_col in df.columns else None

            if pd.isna(part_no) and pd.isna(part_name):
                row_idx += 1
                continue

            no_val = df.iloc[row_idx, no_col] if no_col in df.columns else None
            qty = df.iloc[row_idx, qty_col] if qty_col in df.columns else None

            no_int = to_int_or_none(no_val)
            if no_int is None:
                row_idx += 1
                continue

            row_dict = {
                "PackageId": package_guid,
                "ImagePath": image_filename,
                "Title - TRIM": title,
                "PackageName": package_name,
                "No": no_int,
                "PartNo": str(part_no).strip() if pd.notna(part_no) else None,
                "Part Name And Standard": str(part_name).strip() if pd.notna(part_name) else None,
                "QTY": to_int_or_none(qty),
                # الحقول الأخرى كما اتفقنا سابقاً
                "delete": None,
                "price": None,
                "Description": None,
                "Old Part No.": None,
                "Names and specifications of old parts": None,
                "note": None,
                "is_red": None,
                "is_line": None,
                "is_deleted": None,
                "is_orange": None,
                "internal_notes": None,
            }
            all_rows.append(row_dict)
            row_idx += 1

    return pd.DataFrame(all_rows)


def autosize_columns(path: str):
    """توسيع الأعمدة لكي تظهر القيم كاملة قدر الإمكان."""
    wb = load_workbook(path)
    ws = wb.active

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    wb.save(path)


def main():
    root_dir = os.path.dirname(os.path.abspath(__file__))

    # نحذف فولدر الصور القديم إن وجد لكي لا تتراكم الملفات
    img_root = os.path.join(root_dir, IMAGES_FOLDER_NAME)
    if os.path.isdir(img_root):
        shutil.rmtree(img_root)

    manual_files = find_manual_files(root_dir)

    if not manual_files:
        print("No manual (.xlsx) files found.")
        return

    print("Manual files to be processed:")
    for f in manual_files:
        print(" -", f)

    all_dfs = []

    for file in manual_files:
        try:
            df = extract_manual_to_flat(file, root_dir)
            if not df.empty:
                all_dfs.append(df)
        except Exception as e:
            print("\n⚠ Error processing file:", file)
            print("   Type:", type(e).__name__)
            print("   Message:", e)

    if not all_dfs:
        print("No data generated.")
        return

    final = pd.concat(all_dfs, ignore_index=True)
    output_path = os.path.join(root_dir, OUTPUT_FILENAME)

    try:
        final.to_excel(output_path, index=False, engine="openpyxl")
    except PermissionError:
        print("\n⚠ Cannot write output file. Please close it if open in Excel:")
        print("  ", output_path)
        return

    autosize_columns(output_path)

    print("\nData file generated successfully:")
    print(" ", output_path)
    print("Total rows:", len(final))


if __name__ == "__main__":
    main()
