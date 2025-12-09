import os
from uuid import uuid4
from typing import List, Dict, Any, Tuple, Optional

from openpyxl import load_workbook, Workbook

# ==========================
# إعدادات عامة
# ==========================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# الروت المستهدف الذي يحتوي ملفات الإكسل
ROOT_DIR = os.path.join(SCRIPT_DIR, "TurnKeyFiles2")

# اسم ملف الداتا الناتج (بجانب السكربت أيضاً)
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "packages_data.xlsx")

# كلمات الهيدر التي نبحث عنها (تُقارن بحروف صغيرة)
HEADER_KEYWORDS = ("part number", "description", "qty")

# قيم إكسل الخطأ التي لا يجب اعتبارها أسماء باكجات
EXCLUDED_PACKAGE_TOKENS = {
    "#unknown!",
    "#value!",
    "#div/0!",
    "#ref!",
    "#name?",
    "#null!",
    "#num!",
    "#n/a",
}

# أكواد ألوان ANSI للّوغ (بدون مكتبات إضافية)
RESET = "\033[0m"
CYAN = "\033[36m"
YELLOW = "\033[33m"
GREEN = "\033[32m"
RED = "\033[31m"
MAGENTA = "\033[35m"


def log_info(msg: str) -> None:
    print(f"{CYAN}[INFO]{RESET} {msg}")


def log_warn(msg: str) -> None:
    print(f"{YELLOW}[WARN]{RESET} {msg}")


def log_success(msg: str) -> None:
    print(f"{GREEN}[OK]{RESET}   {msg}")


def log_error(msg: str) -> None:
    print(f"{RED}[ERROR]{RESET} {msg}")


def log_debug(msg: str) -> None:
    print(f"{MAGENTA}[DEBUG]{RESET} {msg}")


# ==========================
# توابع مساعدة عامة
# ==========================

def to_int_or_none(value) -> Optional[int]:
    """
    محاولة تحويل القيمة إلى int وإرجاع None لو فشل التحويل أو كانت القيمة فارغة.
    مشابهة تماماً لما كان يحدث في نسخة pandas.
    """
    if value is None:
        return None
    try:
        text = str(value).strip()
        if not text:
            return None
        # أحياناً تكون القيمة float مثل 3.0
        f = float(text)
        return int(f)
    except (ValueError, TypeError):
        return None


# ==========================
# حساب إحداثيات Y للصفوف
# ==========================

def compute_row_y_map(ws):
    """
    تحسب إحداثيات Y (بالـ EMU) لكل صف:
    top_y[row] = موضع بداية الصف من الأعلى
    bottom_y[row] = موضع نهاية الصف
    نعتمد على ارتفاع الصفوف (إن كان مخصصاً) أو ارتفاع الديفولت من sheet_format.
    """
    sheet_format = ws.sheet_format
    default_height_points = sheet_format.defaultRowHeight or 15  # نقاط
    EMU_PER_POINT = 12700  # ثابت تحويل من نقاط إلى EMU

    top_y = {}
    bottom_y = {}
    acc = 0

    for r in range(1, ws.max_row + 1):
        h = ws.row_dimensions[r].height
        if h is None:
            h = default_height_points
        top_y[r] = acc
        acc += h * EMU_PER_POINT
        bottom_y[r] = acc

    log_info(
        f"Row Y mapping computed using default height {default_height_points} pt "
        f"(~{default_height_points * EMU_PER_POINT:.0f} EMU per row when not custom)."
    )
    return top_y, bottom_y


# ==========================
# منطق استخراج الباكجات
# ==========================

def find_first_header_row(ws) -> int:
    """
    تبحث عن أول سطر يحتوي على أي من الكلمات:
    Part Number / Description / Qty
    في أي عمود من الأعمدة، وترجع رقم السطر (1-based).
    لو لم تجده ترجع 0.
    """
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value:
                continue

            text = str(cell_value).strip().lower()
            if any(keyword in text for keyword in HEADER_KEYWORDS):
                log_info(
                    f"First header row detected at row {row} "
                    f"(col {col}) with value '{cell_value}'"
                )
                return row

    log_warn("No header row found with Part Number / Description / Qty in any column.")
    return 0


def build_packages(ws, top_y, bottom_y) -> List[Dict[str, Any]]:
    """
    تبني قائمة الباكجات اعتماداً على العمود الأول.
    نفس المنطق السابق مع y_start و y_end، لكن بدون صور.
    """
    header_row = find_first_header_row(ws)
    if header_row <= 1:
        log_error("Cannot determine package start row (header row not found or at first row).")
        return []

    start_row = header_row - 1
    log_info(f"Package detection will start from row {start_row} (row above first header).")

    packages: List[Dict[str, Any]] = []
    current_package: Optional[Dict[str, Any]] = None

    for row in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None:
            continue

        text = str(cell_value).strip()
        lower_text = text.lower()

        if any(keyword in lower_text for keyword in HEADER_KEYWORDS):
            continue

        if lower_text in EXCLUDED_PACKAGE_TOKENS:
            log_debug(f"Ignoring error-like value at row {row}: '{text}'")
            continue

        if current_package is not None:
            current_package["end_row"] = row - 1
            current_package["y_start"] = top_y[current_package["start_row"]]
            current_package["y_end"] = bottom_y[current_package["end_row"]]
            packages.append(current_package)

        current_package = {
            "uid": str(uuid4()),
            "name": text,
            "start_row": row,
            "end_row": None,
            "y_start": None,
            "y_end": None,
            "category": None,
        }

    if current_package is not None:
        current_package["end_row"] = ws.max_row
        current_package["y_start"] = top_y[current_package["start_row"]]
        current_package["y_end"] = bottom_y[current_package["end_row"]]
        packages.append(current_package)

    if not packages:
        log_warn("No packages were detected.")
        return []

    log_success(f"{len(packages)} packages detected.")
    return packages


def fill_packages_categories(ws, packages: List[Dict[str, Any]]) -> None:
    """
    ملء Category من العمود F ضمن مدى أسطر كل باكج.
    """
    CATEGORY_COL = 6  # العمود F

    for pkg in packages:
        category = None

        for row in range(pkg["start_row"], pkg["end_row"] + 1):
            cell_value = ws.cell(row=row, column=CATEGORY_COL).value
            if isinstance(cell_value, str):
                text = cell_value.strip()
                if text:
                    category = text
                    break

        pkg["category"] = category
        log_debug(
            f"Package '{pkg['name']}' rows [{pkg['start_row']}-{pkg['end_row']}]: "
            f"Category = '{category}'"
        )


# ==========================
# اكتشاف أعمدة No / PartNo / Desc / QTY
# ==========================

def detect_detail_columns(ws, first_package: Dict[str, Any]) -> Optional[Tuple[int, int, int, int]]:
    """
    نبحث عن عمود رقم السطر (No) بهذه الأولوية:
    1) عمود عنوانه بالضبط "#"
    2) أو عمود عنوانه "no" / "No" / "No." الخ
    ثم نعتبر ما بعده: Part, Desc, QTY.
    """

    start_row = first_package["start_row"]
    end_row = first_package["end_row"]

    FIRST_DATA_COL = 2
    MAX_OFFSET_COLS = 5
    last_col_to_check = FIRST_DATA_COL + MAX_OFFSET_COLS - 1

    candidate_no_cols = []

    for row in range(start_row, end_row + 1):
        for col in range(FIRST_DATA_COL, last_col_to_check + 1):
            cell_value = ws.cell(row=row, column=col).value
            if not isinstance(cell_value, str):
                continue

            raw = cell_value.strip()
            lower = raw.lower()

            is_hash = (raw == "#")
            is_no = (lower in {"no", "no.", "no#", "no:"})

            if not (is_hash or is_no):
                continue

            candidate_no_cols.append((row, col, raw))

    if not candidate_no_cols:
        log_warn(
            "Could not detect detail columns (No / Part Number / Description / QTY) "
            "inside the first package range."
        )
        return None

    row, col_no, header_text = candidate_no_cols[0]

    col_part = col_no + 1
    col_desc = col_no + 2
    col_qty = col_no + 3

    if col_qty > ws.max_column:
        log_warn(
            f"Detected No-like header '{header_text}' at col {col_no} row {row} "
            f"but following columns exceed max_column={ws.max_column}."
        )
        return None

    log_info(
        f"Detail columns detected (row {row}, header '{header_text}'): "
        f"No={col_no}, PartNo={col_part}, Desc={col_desc}, QTY={col_qty}"
    )
    return col_no, col_part, col_desc, col_qty


# ==========================
# فك الدمج العمودي (forward-fill) في أعمدة التفاصيل
# ==========================

def flatten_vertical_merges_in_column(ws, col: int) -> None:
    """
    يفك الدمج العمودي في عمود واحد (مثل عمود No أو QTY):
    - يبحث في ws.merged_cells عن أي range من نوع عمودي في هذا العمود (min_col == max_col == col)
    - يأخذ قيمة الخلية الأولى (أعلى سطر في الدمج)
    - يعمل unmerge للـ range
    - يكتب نفس القيمة في كل الأسطر ضمن هذا الدمج لهذا العمود
    """
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        if min_col == max_col == col and max_row > min_row:
            value = ws.cell(row=min_row, column=col).value
            ws.unmerge_cells(str(merged_range))

            if value is not None and str(value).strip() != "":
                for row in range(min_row, max_row + 1):
                    ws.cell(row=row, column=col).value = value

            log_debug(
                f"Flattened vertical merge in col {col} "
                f"rows [{min_row}-{max_row}] with value '{value}'"
            )


def forward_fill_column_in_range(ws, col: int, start_row: int, end_row: int) -> None:
    last_value = None

    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        value = cell.value

        if value is not None and str(value).strip() != "":
            last_value = value
        else:
            if last_value is not None:
                cell.value = last_value


def find_data_rows_range_for_package(
    ws,
    pkg: Dict[str, Any],
    col_part: int,
    col_desc: int,
) -> Optional[Tuple[int, int]]:
    """
    يحدد نطاق أسطر البيانات (parts) لكل باكج:
    - أي سطر فيه Part Number أو Description نعتبره داتا.
    - يرجع (data_start, data_end) أو None لو لم يوجد داتا.
    """
    data_start = None
    data_end = None

    for row in range(pkg["start_row"], pkg["end_row"] + 1):
        part_val = ws.cell(row=row, column=col_part).value
        desc_val = ws.cell(row=row, column=col_desc).value

        has_part = part_val is not None and str(part_val).strip() != ""
        has_desc = desc_val is not None and str(desc_val).strip() != ""

        if has_part or has_desc:
            if data_start is None:
                data_start = row
            data_end = row

    if data_start is None or data_end is None:
        return None

    return data_start, data_end


def normalize_merged_detail_cells_for_all_packages(
    ws,
    packages: List[Dict[str, Any]],
    col_no: int,
    col_part: int,
    col_desc: int,
    col_qty: int,
) -> Dict[int, Tuple[int, int]]:
    """
    تعوّض الدمج العمودي داخل أعمدة تفاصيل القطع ضمن مجال أسطر الداتا لكل باكج.
    ترجع dict يربط package_index → (data_start, data_end).
    """

    flatten_vertical_merges_in_column(ws, col_no)
    flatten_vertical_merges_in_column(ws, col_qty)

    data_ranges: Dict[int, Tuple[int, int]] = {}

    for idx, pkg in enumerate(packages):
        res = find_data_rows_range_for_package(ws, pkg, col_part, col_desc)
        if res is None:
            log_warn(f"No data rows found for package '{pkg['name']}'.")
            continue

        data_start, data_end = res
        data_ranges[idx] = (data_start, data_end)

        forward_fill_column_in_range(ws, col_no, data_start, data_end)
        forward_fill_column_in_range(ws, col_qty, data_start, data_end)

        log_debug(
            f"Forward-filled No/QTY for package '{pkg['name']}' "
            f"rows [{data_start}-{data_end}]."
        )

    return data_ranges


# ==========================
# معالجة ملف واحد واستخراج أسطر القطع
# ==========================

def process_workbook(path: str) -> List[tuple]:
    """
    تعالج ملف إكسل واحد:
    - تبني الباكجات + الفئات + الأعمدة التفصيلية + فك الدمج
    - تبني أسطر القطع لكل باكج:
      (PackageId, TitleTrim, PackageName,
       No, PartNo, PartNameAndStandard, QTY, Category)
    """
    basename = os.path.basename(path)
    title_trim = os.path.splitext(basename)[0].strip()

    log_info(f"Opening workbook: {basename}")

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        log_error(f"Failed to open '{basename}': {e}")
        return []

    ws = wb.active

    top_y, bottom_y = compute_row_y_map(ws)

    packages = build_packages(ws, top_y, bottom_y)
    if not packages:
        log_warn(f"No packages found in '{basename}'.")
        return []

    fill_packages_categories(ws, packages)

    detail_cols = detect_detail_columns(ws, packages[0])
    rows_for_excel: List[tuple] = []

    if detail_cols is None:
        log_warn(
            f"Detail columns could not be detected in '{basename}'. "
            f"Only package-level rows will be emitted without part details."
        )
        # fallback: سطر واحد لكل باكج بدون تفاصيل قطع
        for pkg in packages:
            uid = pkg.get("uid")
            pkg_name = pkg["name"]
            category = pkg.get("category") or ""
            rows_for_excel.append(
                (uid, title_trim, pkg_name, "", "", "", "", category)
            )
        return rows_for_excel

    col_no, col_part, col_desc, col_qty = detail_cols

    data_ranges_by_pkg_index = normalize_merged_detail_cells_for_all_packages(
        ws,
        packages,
        col_no,
        col_part,
        col_desc,
        col_qty,
    )

    # نمر على كل باكج ونبني أسطر القطع
    for idx, pkg in enumerate(packages):
        if idx not in data_ranges_by_pkg_index:
            # باكج بدون داتا حقيقية
            continue

        data_start, data_end = data_ranges_by_pkg_index[idx]

        uid = pkg.get("uid")
        pkg_name = pkg["name"]
        category = pkg.get("category") or ""

        for row in range(data_start, data_end + 1):
            no_val = ws.cell(row=row, column=col_no).value
            part_val = ws.cell(row=row, column=col_part).value
            desc_val = ws.cell(row=row, column=col_desc).value
            qty_val = ws.cell(row=row, column=col_qty).value

            has_part = part_val is not None and str(part_val).strip() != ""
            has_desc = desc_val is not None and str(desc_val).strip() != ""
            if not (has_part or has_desc):
                continue

            no_int = to_int_or_none(no_val)
            if no_int is None:
                continue

            qty_int = to_int_or_none(qty_val)

            part_str = str(part_val).strip() if part_val is not None else ""
            desc_str = str(desc_val).strip() if desc_val is not None else ""

            rows_for_excel.append(
                (
                    uid,            # PackageId
                    title_trim,     # Title - TRIM
                    pkg_name,       # PackageName
                    no_int,         # No
                    part_str,       # PartNo
                    desc_str,       # Part Name And Standard
                    qty_int,        # QTY
                    category,       # Category
                )
            )

    return rows_for_excel


# ==========================
# الدالة الرئيسية
# ==========================

def main():
    """
    - يعالج كل ملف .xlsx في ROOT_DIR وفي المجلدات الفرعية داخله.
    - يبني ملف إكسل جديد packages_data.xlsx
      يحتوي أسطر القطع مع تكرار بيانات الباكج لكل سطر.
    - لا يتعامل مع الصور إطلاقاً.
    """
    # نجمع كل ملفات .xlsx في ROOT_DIR وفي كل المجلدات الفرعية
    xlsx_files: List[str] = []
    for dirpath, _, filenames in os.walk(ROOT_DIR):
        for name in filenames:
            lower = name.lower()
            if not lower.endswith(".xlsx"):
                continue
            if name.startswith("~$"):
                continue
            xlsx_files.append(os.path.join(dirpath, name))

    if not xlsx_files:
        log_error(f"No .xlsx files found under ROOT_DIR: {ROOT_DIR}")
        return

    log_info(f"Found {len(xlsx_files)} .xlsx file(s) under ROOT_DIR: {ROOT_DIR}")
    for path in xlsx_files:
        rel = os.path.relpath(path, ROOT_DIR)
        log_debug(f"- {rel}")

    all_rows: List[tuple] = []

    for full_path in xlsx_files:
        print()
        rel = os.path.relpath(full_path, ROOT_DIR)
        print(f"{MAGENTA}========== Processing file: {rel} =========={RESET}")
        rows = process_workbook(full_path)
        all_rows.extend(rows)

    if not all_rows:
        log_warn("No rows were collected. Excel data file will not be created.")
        return

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "packages"

    # عرض الأعمدة (بدون أعمدة صور)
    ws_out.column_dimensions["A"].width = 40   # PackageId
    ws_out.column_dimensions["B"].width = 30   # Title - TRIM
    ws_out.column_dimensions["C"].width = 30   # PackageName
    ws_out.column_dimensions["D"].width = 5    # No
    ws_out.column_dimensions["E"].width = 18   # PartNo
    ws_out.column_dimensions["F"].width = 30   # Part Name And Standard
    ws_out.column_dimensions["G"].width = 6    # QTY
    ws_out.column_dimensions["H"].width = 20   # Category

    header = [
        "PackageId",
        "Title - TRIM",
        "PackageName",
        "No",
        "PartNo",
        "Part Name And Standard",
        "QTY",
        "Category",
        "delete",
        "price",
        "Description",
        "Old Part No.",
        "Names and specifications of old parts",
        "note",
        "is_red",
        "is_line",
        "is_deleted",
        "is_orange",
        "is_pink",
        "is_yellow",
        "internal_notes",
    ]
    ws_out.append(header)

    num_cols = len(header)

    row_idx = 1
    last_pkg_key = None

    for row_data in all_rows:
        (
            uid,
            title_trim,
            pkg_name,
            no_val,
            part_no,
            part_name_std,
            qty_val,
            category,
        ) = row_data

        pkg_key = (uid, title_trim, pkg_name)

        if last_pkg_key is not None and pkg_key != last_pkg_key:
            row_idx += 1
            ws_out.append([""] * num_cols)

        row_idx += 1
        ws_out.append([
            uid,           # PackageId
            title_trim,    # Title - TRIM
            pkg_name,      # PackageName
            no_val,        # No
            part_no,       # PartNo
            part_name_std, # Part Name And Standard
            qty_val,       # QTY
            category,      # Category
            "",            # delete
            "",            # price
            "",            # Description
            "",            # Old Part No.
            "",            # Names and specifications of old parts
            "",            # note
            "",            # is_red
            "",            # is_line
            "",            # is_deleted
            "",            # is_orange
            "",            # is_pink
            "",            # is_yellow
            "",            # internal_notes
        ])

        last_pkg_key = pkg_key

    wb_out.save(OUTPUT_EXCEL)
    log_success(f"Data Excel file created: '{OUTPUT_EXCEL}'")


if __name__ == "__main__":
    main()
